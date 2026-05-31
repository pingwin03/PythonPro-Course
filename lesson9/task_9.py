# Prosty arkusz kalkulacyjny: Używając openpyxl , stwórz plik finanse.xlsx . W
# pierwszej kolumnie umieść nazwy wydatków (np. "Czynsz", "Jedzenie"), a w drugiej ich
# wartości. W komórce poniżej wartości oblicz i wstaw sumę wszystkich wydatków, używając
# formuły Excela (np. =SUM(B1:B2) )

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Finanse"

wydatki = [
    ("Czynsz", 1500),
    ("Jedzenie", 650),
]

ws["A1"] = "Nazwa"
ws["B1"] = "Wartość"

for row, (nazwa, kwota) in enumerate(wydatki, start=2):
    ws.cell(row=row, column=1, value=nazwa)
    ws.cell(row=row, column=2, value=kwota)

ostatni_wiersz = len(wydatki) + 2
ws.cell(row=ostatni_wiersz, column=1, value="Suma")
ws.cell(row=ostatni_wiersz, column=2, value=f"=SUM(B2:B{ostatni_wiersz-1})")

wb.save("finanse.xlsx")